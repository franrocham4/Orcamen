"""
ProcessadorExcel: Reads all sheets and rows from an Excel file and saves to the database.
"""
import logging
import os
from datetime import datetime, timezone

import openpyxl

logger = logging.getLogger(__name__)


def _cell_value(value):
    """Convert cell value to JSON-serializable type."""
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y %H:%M') if value.hour or value.minute else value.strftime('%d/%m/%Y')
    if value is None:
        return ''
    return str(value) if not isinstance(value, (int, float, bool, str)) else value


class ProcessadorExcel:
    """Reads all sheets and data from an Excel file."""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.filename = os.path.basename(filepath)

    def processar(self, sync_history=None):
        """
        Main entry point: process the Excel file and persist to DB.
        Returns a dict with stats.
        """
        from apps.core.models import ExcelFile, ExcelSheet, ExcelRow, SyncHistory

        logger.info(f'Processing Excel file: {self.filepath}')

        if not os.path.exists(self.filepath):
            msg = f'File not found: {self.filepath}'
            logger.error(msg)
            if sync_history:
                sync_history.status = 'error'
                sync_history.error_message = msg
                sync_history.save()
            return {'success': False, 'error': msg}

        try:
            # keep_vba=True is required for .xlsm files; data_only=True reads computed cell values
            wb = openpyxl.load_workbook(self.filepath, keep_vba=True, data_only=True)
        except Exception as e:
            msg = f'Could not open workbook: {e}'
            logger.error(msg)
            if sync_history:
                sync_history.status = 'error'
                sync_history.error_message = msg
                sync_history.save()
            return {'success': False, 'error': msg}

        # Get or create ExcelFile record
        file_stat = os.stat(self.filepath)
        excel_file, _ = ExcelFile.objects.update_or_create(
            filepath=self.filepath,
            defaults={
                'filename': self.filename,
                'file_size': file_stat.st_size,
                'last_modified': datetime.fromtimestamp(file_stat.st_mtime, tz=timezone.utc),
                'is_active': True,
            }
        )

        if sync_history:
            sync_history.excel_file = excel_file
            sync_history.save()

        total_sheets = 0
        total_rows = 0

        for sheet_index, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows_data, headers = self._extract_sheet_data(ws)

            # Delete old rows and re-create
            sheet_obj, _ = ExcelSheet.objects.update_or_create(
                excel_file=excel_file,
                name=sheet_name,
                defaults={
                    'sheet_index': sheet_index,
                    'row_count': len(rows_data),
                    'col_count': ws.max_column or 0,
                    'headers': headers,
                }
            )

            # Bulk replace rows
            ExcelRow.objects.filter(sheet=sheet_obj).delete()
            row_objects = [
                ExcelRow(sheet=sheet_obj, row_number=i, data=row)
                for i, row in enumerate(rows_data)
            ]
            ExcelRow.objects.bulk_create(row_objects, batch_size=500)

            total_sheets += 1
            total_rows += len(rows_data)
            logger.info(f'  Sheet "{sheet_name}": {len(rows_data)} rows')

        logger.info(f'Done: {total_sheets} sheets, {total_rows} rows')
        return {
            'success': True,
            'sheets': total_sheets,
            'rows': total_rows,
            'filename': self.filename,
        }

    def _extract_sheet_data(self, ws):
        """
        Extract all non-empty rows from a worksheet.
        Returns (list_of_dicts, list_of_headers).
        """
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []

        # Find the header row: first row that has at least 2 non-None values
        header_row_idx = 0
        for i, row in enumerate(all_rows):
            non_none = [v for v in row if v is not None]
            if len(non_none) >= 2:
                header_row_idx = i
                break

        raw_headers = all_rows[header_row_idx]
        headers = [_cell_value(h) if h is not None else f'Col{i+1}' for i, h in enumerate(raw_headers)]

        # Deduplicate headers
        seen = {}
        unique_headers = []
        for h in headers:
            h_str = str(h).strip() or 'Col'
            if h_str in seen:
                seen[h_str] += 1
                unique_headers.append(f'{h_str}_{seen[h_str]}')
            else:
                seen[h_str] = 0
                unique_headers.append(h_str)

        rows_data = []
        for row in all_rows[header_row_idx + 1:]:
            # Skip completely empty rows
            if all(v is None for v in row):
                continue
            row_dict = {unique_headers[i]: _cell_value(v) for i, v in enumerate(row) if i < len(unique_headers)}
            rows_data.append(row_dict)

        return rows_data, unique_headers
